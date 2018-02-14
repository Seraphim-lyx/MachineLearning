from openpyxl import load_workbook
import copy
import math
import random


class linearRegression(object):
    """docstring for LenearRegression"""

    def __init__(self):
        self.theta = [0, 0, 0, 0, 0]
        self.mean = [0, 0, 0, 0]
        self.sd = [0, 0, 0, 0]
        self.dataSet = []
        self.trainSet = []
        self.testSet = []

    def loadData(self, file, sheet):
        """
        loading raw data
        """
        wb = load_workbook(filename=file)
        self.ws = wb[sheet]
        # self.setFactor()
        for row in self.ws.rows:
            r = []
            r.append(1)
            r.append(row[2].value)
            r.append(row[3].value)
            r.append(row[4].value)
            r.append(row[5].value)
            r.append(row[1].value)
            self.dataSet.append(r)

        self.max_row = len(self.dataSet)

    def learning(self):
        """
        iterate each attributes for gradient descent
        """
        lastCF = 0
        CF = self.costFunction(self.theta)
        while abs(lastCF - CF) > 0.1:
            self.batchGD(CF)
            lastCF = CF
            CF = self.costFunction(self.theta)
            print(self.theta, lastCF, CF)

    def batchGD(self, minCF):
        """
        """
        # print(thetaList)
        localTheta = self.theta[:]
        total = 0
        alpha = 0.0001
        for j in range(len(self.theta)):
            for row in self.dataSet[:40000]:
                # row_ = self.reOrderValue(row)
                err = self.error(row, localTheta)
                total += err * row[j]
            self.theta[j] = self.theta[j] - (total / self.max_row) * alpha

    def costFunction(self, tl):
        total = 0

        for row in self.dataSet[:40000]:
            err = self.error(row, tl)
            total += pow(err, 2)

        return total / (2 * self.max_row)

    def error(self, row, t):
        """
        calculate the distance between
        computing result and real result
        """

        y = t[0] * row[0] + \
            t[1] * row[1] + \
            t[2] * row[2] + \
            t[3] * row[3] + \
            t[4] * row[4]
        return y - row[5]

    def reOrderValue(self, row):
        r = []
        r.append(1)  # const
        r.append(row[2].value)  # temp
        r.append(row[3].value)  # humidity
        r.append(row[4].value)  # wind speed
        r.append(row[5].value)  # vapour pressure
        r.append(row[1].value)  # MW
        return r

    def test(self):
        # print(self.theta)
        total = 0
        for row in self.dataSet[40000:]:
            y = self.theta[0] * row[0] + \
                self.theta[1] * row[1] + \
                self.theta[2] * row[2] + \
                self.theta[3] * row[3] + \
                self.theta[4] * row[4]
            total += abs(y - row[5])
        print(total/self.max_row)


l = linearRegression()
l.loadData('data.xlsx', 'Sheet1')

l.learning()
l.test()
