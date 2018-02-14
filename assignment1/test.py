from openpyxl import load_workbook

import math


wb = load_workbook(filename='data1.xlsx')
ws = wb['Sheet1']

for row in ws.iter_rows(min_row=5, max_row=10):
    print(row)


[1.2450348399099198, 140.24729686717893, 3.5610502974044036, -14.608456142496287, 51.441167706290436] 72763.55151148335 72763.45160530935
