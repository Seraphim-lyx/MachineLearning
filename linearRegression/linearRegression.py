from openpyxl import load_workbook
import random


class linearRegression(object):
    """docstring for LenearRegression"""

    def __init__(self):
        self.theta = [0, 0, 0, 0, 0]
        self.mean = [0, 0, 0, 0]
        self.sd = [0, 0, 0, 0]
        self.dataSet = []
        self.trainSet = []
        self.testSet = []

    def loadData(self, file, sheet):
        """
        loading raw data
        """
        wb = load_workbook(filename=file)
        self.ws = wb[sheet]
        for row in self.ws.rows:
            r = []
            r.append(1)
            r.append(row[2].value)
            r.append(row[3].value)
            r.append(row[4].value)
            r.append(row[5].value)
            r.append(row[1].value)
            self.dataSet.append(r)

        for i in self.dataSet:
            if random.random() > 0.8:
                self.testSet.append(i)
            else:
                self.trainSet.append(i)

        self.max_row = len(self.dataSet)

    def learning(self):
        """
        iterate each attributes for gradient descent
        """
        lastCF = 0
        CF = self.costFunction(self.theta)
        while abs(lastCF - CF) > 0.01:
            self.batchGD(CF)
            lastCF = CF
            CF = self.costFunction(self.theta)
            print(self.theta, lastCF, CF)

    def batchGD(self, minCF):
        """
        """
        localTheta = self.theta[:]
        total = 0
        alpha = 0.0002

        for j in range(len(self.theta)):
            for row in self.trainSet:
                err = self.error(row, localTheta)
                total += err * row[j]
            self.theta[j] = self.theta[j] - \
                (total / self.max_row) * alpha  # deviation

    def costFunction(self, tl):
        total = 0

        for row in self.trainSet:
            err = self.error(row, tl)
            total += pow(err, 2)

        return total / (2 * self.max_row)

    def error(self, row, t):
        """
        calculate the distance between
        computing result and real result
        """

        y = t[0] * row[0] + \
            t[1] * row[1] + \
            t[2] * row[2] + \
            t[3] * row[3] + \
            t[4] * row[4]
        return y - row[5]

    def test(self):
        total = 0
        for row in self.testSet:
            y = self.theta[0] * row[0] + \
                self.theta[1] * row[1] + \
                self.theta[2] * row[2] + \
                self.theta[3] * row[3] + \
                self.theta[4] * row[4]
            total += pow(y - row[5], 2)
        print(total / (2 * len(self.testSet)))


l = linearRegression()
l.loadData('data.xlsx', 'Sheet1')
l.learning()
# l.theta = [70.61729996, 141.79027405, 1.61074663, -16.15458898, 56.48189629]
l.test()
